
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from functools import lru_cache
from io import BytesIO
import re

import pandas as pd
from flask import Blueprint, render_template, request, current_app, send_file, url_for, redirect
from openpyxl.styles import Alignment

from db import get_db, get_teams_db
from po_security import (
    current_request_next_url,
    filter_records_for_po_access,
    get_current_po_access,
    po_pin_security_enabled,
)
from routes.po import fetch_po_records, get_supabase_client
from services.reference_data import get_candidate_lookup_names, get_teams_reference
from services.team_management import (
    get_management_snapshot,
    get_team_management_directory,
    mongo_normalized_text,
    normalize_lookup_text,
    normalize_person_name,
)

candidates_bp = Blueprint('candidates', __name__)

PO_MATCH_THRESHOLD = 0.8
PO_RECORDS_CACHE_KEY = "expert_activity_po_records_v2"
EXPERT_ACTIVITY_CACHE_VERSION = "v6"

SUBJECT_MONTH_MAP = {
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
    'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
    'sep': '09', 'sept': '09', 'oct': '10', 'nov': '11', 'dec': '12',
    'january': '01', 'february': '02', 'march': '03', 'april': '04',
    'june': '06', 'july': '07', 'august': '08', 'september': '09',
    'october': '10', 'november': '11', 'december': '12',
}

SUBJECT_MONTH_TOKEN_PATTERN = (
    r'Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|'
    r'Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|'
    r'Nov(?:ember)?|Dec(?:ember)?'
)

SUBJECT_WEEKDAY_TOKEN_PATTERN = (
    r'Mon(?:day)?|Tue(?:s(?:day)?)?|Wed(?:nesday)?|Thu(?:rs(?:day)?)?|'
    r'Fri(?:day)?|Sat(?:urday)?|Sun(?:day)?'
)


def get_team_options():
    reference = get_teams_reference()
    return reference["teams_list"], reference["all_experts"], reference["teams_map"]


def expert_activity_cache_key(month_s, year_s, status_f, team_f, expert_f, include_all_candidates, exclude_rounds):
    rounds_key = exclude_rounds
    if isinstance(exclude_rounds, (list, tuple, set)):
        rounds_key = ",".join(sorted(str(value).strip() for value in exclude_rounds if str(value).strip()))

    return ":".join([
        "candidate-activity",
        EXPERT_ACTIVITY_CACHE_VERSION,
        str(month_s or "all"),
        str(year_s or "all"),
        str(status_f or "all"),
        str(team_f or "all"),
        str(expert_f or "all"),
        "full" if include_all_candidates else "active",
        str(rounds_key or "none"),
    ])


def normalize_expert_match_name(value):
    cleaned = " ".join(str(value or "").replace("\xa0", " ").split())
    if not cleaned:
        return ""

    if "@" in cleaned:
        cleaned = cleaned.split("@", 1)[0]

    cleaned = re.sub(r"[._-]+", " ", cleaned)
    cleaned = re.sub(r"[^A-Za-z0-9\s]", " ", cleaned)
    cleaned = " ".join(cleaned.split())
    if not cleaned:
        return ""

    return " ".join(normalize_person_name(cleaned).lower().split())


def get_name_match_score(left, right):
    if not left or not right:
        return 0.0

    if left == right:
        return 1.0

    left_sorted = " ".join(sorted(left.split()))
    right_sorted = " ".join(sorted(right.split()))

    return max(
        SequenceMatcher(None, left, right).ratio(),
        SequenceMatcher(None, left_sorted, right_sorted).ratio(),
    )


def get_selected_po_month_key(month_s, year_s):
    selected_month = parse_selected_activity_month(month_s, year_s)
    return selected_month.strftime("%Y-%m") if selected_month else ""


@lru_cache(maxsize=256)
def parse_selected_activity_month(month_s, year_s):
    month_text = str(month_s or "").strip().title()
    year_text = str(year_s or "").strip()

    if not month_text or not year_text:
        return None

    for fmt in ("%b %Y", "%B %Y"):
        try:
            return datetime.strptime(f"{month_text} {year_text}", fmt)
        except ValueError:
            continue

    return None


@lru_cache(maxsize=4096)
def extract_interview_date_candidates_from_subject(subject):
    if not subject:
        return ()

    normalized_subject = " ".join(
        str(subject)
        .replace("\xa0", " ")
        .replace("\u202f", " ")
        .split()
    )
    if not normalized_subject:
        return ()

    candidates = []
    seen = set()

    def add_candidate(year_value, month_value, day_value):
        try:
            candidate = datetime(
                int(year_value),
                int(month_value),
                int(day_value),
            ).strftime("%Y-%m-%d")
        except (TypeError, ValueError):
            return

        if candidate not in seen:
            seen.add(candidate)
            candidates.append(candidate)

    month_first_pattern = re.compile(
        rf'\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?({SUBJECT_MONTH_TOKEN_PATTERN})\.?\s*(?:,)?\s*(\d{{1,2}})(?:st|nd|rd|th)?\s*(?:,)?\s*(\d{{4}})\b',
        re.IGNORECASE,
    )
    day_first_pattern = re.compile(
        rf'\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?(\d{{1,2}})(?:st|nd|rd|th)?\s+({SUBJECT_MONTH_TOKEN_PATTERN})\.?\s*(?:,)?\s*(\d{{4}})\b',
        re.IGNORECASE,
    )
    iso_pattern = re.compile(r'\b(\d{4})[/-](\d{1,2})[/-](\d{1,2})\b')
    numeric_pattern = re.compile(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b')
    numeric_short_year_pattern = re.compile(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{2})\b')

    for month_str, day, year in month_first_pattern.findall(normalized_subject):
        month = SUBJECT_MONTH_MAP.get(month_str.lower().rstrip('.'))
        if month:
            add_candidate(year, month, day)

    for day, month_str, year in day_first_pattern.findall(normalized_subject):
        month = SUBJECT_MONTH_MAP.get(month_str.lower().rstrip('.'))
        if month:
            add_candidate(year, month, day)

    for year, month, day in iso_pattern.findall(normalized_subject):
        add_candidate(year, month, day)

    for first, second, year in numeric_pattern.findall(normalized_subject):
        first_num = int(first)
        second_num = int(second)

        if 1 <= first_num <= 12 and 1 <= second_num <= 31:
            add_candidate(year, first_num, second_num)
        if 1 <= first_num <= 31 and 1 <= second_num <= 12:
            add_candidate(year, second_num, first_num)

    for first, second, year in numeric_short_year_pattern.findall(normalized_subject):
        expanded_year = f"20{year}"
        first_num = int(first)
        second_num = int(second)

        if 1 <= first_num <= 12 and 1 <= second_num <= 31:
            add_candidate(expanded_year, first_num, second_num)
        if 1 <= first_num <= 31 and 1 <= second_num <= 12:
            add_candidate(expanded_year, second_num, first_num)

    return tuple(candidates)


def parse_interview_date_from_subject(subject, reference_date=None):
    """
    Extract interview date from subject line and return YYYY-MM-DD.
    """
    candidates = extract_interview_date_candidates_from_subject(subject)
    if not candidates:
        candidates = ()

    reference_value = str(reference_date or "").strip()[:10]
    try:
        reference_dt = datetime.strptime(reference_value, "%Y-%m-%d")
    except ValueError:
        reference_dt = None

    if not candidates and reference_dt:
        normalized_subject = " ".join(
            str(subject or "")
            .replace("\xa0", " ")
            .replace("\u202f", " ")
            .split()
        )
        inferred_candidates = []
        seen = set()

        def add_inferred_candidate(year_value, month_value, day_value):
            try:
                candidate = datetime(
                    int(year_value),
                    int(month_value),
                    int(day_value),
                ).strftime("%Y-%m-%d")
            except (TypeError, ValueError):
                return

            if candidate not in seen:
                seen.add(candidate)
                inferred_candidates.append(candidate)

        month_first_partial_year_pattern = re.compile(
            rf'\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?({SUBJECT_MONTH_TOKEN_PATTERN})\.?\s*(\d{{1,2}})(?:st|nd|rd|th)?\s*(?:,)?\s*(\d{{3}})\b',
            re.IGNORECASE,
        )
        day_first_partial_year_pattern = re.compile(
            rf'\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?(\d{{1,2}})(?:st|nd|rd|th)?\s+({SUBJECT_MONTH_TOKEN_PATTERN})\.?\s*(?:,)?\s*(\d{{3}})\b',
            re.IGNORECASE,
        )
        month_first_no_year_pattern = re.compile(
            rf'\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?({SUBJECT_MONTH_TOKEN_PATTERN})\.?\s*(\d{{1,2}})(?:st|nd|rd|th)?\b',
            re.IGNORECASE,
        )
        day_first_no_year_pattern = re.compile(
            rf'\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?(\d{{1,2}})(?:st|nd|rd|th)?\s+({SUBJECT_MONTH_TOKEN_PATTERN})\.?\b',
            re.IGNORECASE,
        )

        for month_str, day, year_fragment in month_first_partial_year_pattern.findall(normalized_subject):
            if str(reference_dt.year).startswith(year_fragment):
                month = SUBJECT_MONTH_MAP.get(month_str.lower().rstrip('.'))
                if month:
                    add_inferred_candidate(reference_dt.year, month, day)

        for day, month_str, year_fragment in day_first_partial_year_pattern.findall(normalized_subject):
            if str(reference_dt.year).startswith(year_fragment):
                month = SUBJECT_MONTH_MAP.get(month_str.lower().rstrip('.'))
                if month:
                    add_inferred_candidate(reference_dt.year, month, day)

        if not inferred_candidates:
            for month_str, day in month_first_no_year_pattern.findall(normalized_subject):
                month = SUBJECT_MONTH_MAP.get(month_str.lower().rstrip('.'))
                if month:
                    add_inferred_candidate(reference_dt.year, month, day)

            for day, month_str in day_first_no_year_pattern.findall(normalized_subject):
                month = SUBJECT_MONTH_MAP.get(month_str.lower().rstrip('.'))
                if month:
                    add_inferred_candidate(reference_dt.year, month, day)

        candidates = tuple(inferred_candidates)

    if not candidates:
        return None

    if len(candidates) == 1 or not reference_dt:
        return candidates[0]

    def candidate_distance(candidate_value):
        try:
            candidate_dt = datetime.strptime(candidate_value, "%Y-%m-%d")
        except ValueError:
            return (float("inf"), candidate_value)
        return (abs((candidate_dt - reference_dt).days), candidate_value)

    return min(candidates, key=candidate_distance)


def build_selected_month_subject_regex(month_s, year_s):
    selected_month = parse_selected_activity_month(month_s, year_s)
    if not selected_month:
        return None

    month_tokens = [selected_month.strftime("%b"), selected_month.strftime("%B")]
    if selected_month.strftime("%b") == "Sep":
        month_tokens.append("Sept")

    month_pattern = "|".join(re.escape(token) for token in month_tokens)
    numeric_month_pattern = rf"(?:0?{selected_month.month}|{selected_month.strftime('%m')})"
    return (
        rf"(?:"
        rf"\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?(?:{month_pattern})\.?\s*\d{{1,2}}(?:st|nd|rd|th)?\s*(?:,)?\s*{selected_month.year}\b"
        rf"|"
        rf"\b(?:(?:{SUBJECT_WEEKDAY_TOKEN_PATTERN})\.?,?\s+)?\d{{1,2}}(?:st|nd|rd|th)?\s+(?:{month_pattern})\.?\s*(?:,)?\s*{selected_month.year}\b"
        rf"|"
        rf"\b{selected_month.year}[/-]{numeric_month_pattern}[/-]\d{{1,2}}\b"
        rf"|"
        rf"\b{numeric_month_pattern}[/-]\d{{1,2}}[/-]{selected_month.year}\b"
        rf"|"
        rf"\b\d{{1,2}}[/-]{numeric_month_pattern}[/-]{selected_month.year}\b"
        rf")"
    )


def build_selected_month_received_date_match(month_s, year_s):
    selected_month = parse_selected_activity_month(month_s, year_s)
    if not selected_month:
        return None

    if selected_month.month == 12:
        next_month = selected_month.replace(year=selected_month.year + 1, month=1)
    else:
        next_month = selected_month.replace(month=selected_month.month + 1)

    return {
        "$gte": selected_month.strftime("%Y-%m-01T00:00:00"),
        "$lt": next_month.strftime("%Y-%m-01T00:00:00"),
    }


def interview_matches_selected_month(interview_detail, selected_month_key):
    if not selected_month_key:
        return True

    subject_date = parse_interview_date_from_subject(
        interview_detail.get("Subject", ""),
        reference_date=interview_detail.get("Date", ""),
    )
    if not subject_date:
        received_value = str(interview_detail.get("Date", "") or "").strip()
        return received_value.startswith(selected_month_key)

    return subject_date.startswith(selected_month_key)


def get_cached_po_records():
    cache = getattr(current_app, "cache", None)
    if cache:
        cached_records = cache.get(PO_RECORDS_CACHE_KEY)
        if cached_records is not None:
            return cached_records

    records = fetch_po_records(get_supabase_client())

    if cache:
        cache.set(PO_RECORDS_CACHE_KEY, records, timeout=300)

    return records


def build_po_counts_by_expert(month_s, year_s, expert_names):
    normalized_experts = {
        normalize_lookup_text(name): normalize_expert_match_name(name)
        for name in expert_names
        if normalize_lookup_text(name)
    }
    if not normalized_experts:
        return {}

    if po_pin_security_enabled() and not get_current_po_access():
        return {}

    try:
        po_records = filter_records_for_po_access(get_cached_po_records(), get_current_po_access())
    except Exception as exc:
        current_app.logger.warning("Unable to load PO counts for expert activity: %s", exc)
        return {}

    selected_po_month = get_selected_po_month_key(month_s, year_s)
    po_counts = defaultdict(int)
    for record in po_records:
        if selected_po_month and record.get("month_key") != selected_po_month:
            continue

        expert_name = record.get("expert_name") or record.get("interview_support_by")
        normalized_name = normalize_expert_match_name(expert_name)
        if not normalized_name or normalized_name == "unassigned":
            continue

        po_counts[normalized_name] += 1

    if not po_counts:
        return {}

    matched_counts = {}
    unmatched_experts = {}
    remaining_po_names = set(po_counts.keys())

    for expert_key, normalized_name in normalized_experts.items():
        if normalized_name and normalized_name in po_counts:
            matched_counts[expert_key] = po_counts[normalized_name]
            remaining_po_names.discard(normalized_name)
        else:
            unmatched_experts[expert_key] = normalized_name

    candidate_matches = []
    for expert_key, normalized_name in unmatched_experts.items():
        if not normalized_name:
            continue

        for po_name in remaining_po_names:
            score = get_name_match_score(normalized_name, po_name)
            if score >= PO_MATCH_THRESHOLD:
                candidate_matches.append((score, expert_key, po_name))

    for _, expert_key, po_name in sorted(candidate_matches, key=lambda item: item[0], reverse=True):
        if expert_key in matched_counts or po_name not in remaining_po_names:
            continue

        matched_counts[expert_key] = po_counts[po_name]
        remaining_po_names.remove(po_name)

    return matched_counts

@candidates_bp.route('/', methods=['GET'])
def search():
    query_name = request.args.get('q', '')
    results = []

    if query_name:
        db = get_db()
        # Case insensitive regex match for Candidate Name
        mongo_query = {
            "Candidate Name": {"$regex": query_name, "$options": 'i'}
        }

        # Querying candidateDetails collection with projection
        cursor = db.candidateDetails.find(
            mongo_query,
            {
                "Candidate Name": 1,
                "workflowStatus": 1,
                "Technology": 1,
                "_id": 1
            }
        ).limit(50)
        results = list(cursor)

    return render_template('search.html', query=query_name, results=results)


@candidates_bp.route('/lookup', methods=['GET'])
def candidate_lookup():
    """
    Candidate interview lookup - shows detailed interview statistics
    for a specific candidate from the taskBody collection.
    """
    db = get_db()
    candidate_name = request.args.get('name', '').strip()
    candidate_name_key = normalize_lookup_text(candidate_name)

    # Get list of all unique candidate names for autocomplete/dropdown - OPTIMIZED with limit
    all_candidates = get_candidate_lookup_names(limit=500)

    candidate_data = None
    interview_records = []

    if candidate_name_key:
        cache_key = f"candidate-lookup:summary:{candidate_name_key}"
        cache = getattr(current_app, "cache", None)
        cached = cache.get(cache_key) if cache else None

        if cached is None:
            pipeline = [
                {
                    "$match": {
                        "$expr": {
                            "$eq": [
                                mongo_normalized_text("Candidate Name"),
                                candidate_name_key,
                            ]
                        }
                    }
                },
                {
                    "$facet": {
                        "totalInterviews": [
                            {"$count": "count"}
                        ],
                        "byRound": [
                            {"$group": {"_id": "$actualRound", "count": {"$sum": 1}}},
                            {"$sort": {"count": -1}}
                        ],
                        "byStatus": [
                            {"$group": {"_id": "$status", "count": {"$sum": 1}}},
                            {"$sort": {"count": -1}}
                        ],
                        "byExpert": [
                            {"$group": {"_id": "$assignedTo", "count": {"$sum": 1}}},
                            {"$sort": {"count": -1}}
                        ],
                        "timeline": [
                            {"$sort": {"receivedDateTime": -1}},
                            {"$limit": 50},
                            {"$project": {
                                "subject": 1,
                                "actualRound": 1,
                                "status": 1,
                                "assignedTo": 1,
                                "receivedDateTime": 1,
                                "scheduledDateTime": 1
                            }}
                        ]
                    }
                }
            ]

            results = list(db.taskBody.aggregate(pipeline))

            payload = {
                "candidate_data": None,
                "interview_records": [],
            }

            if results:
                data = results[0]
                total = data["totalInterviews"][0]["count"] if data["totalInterviews"] else 0

                by_round = []
                for r in data["byRound"]:
                    round_name = r['_id'] if r['_id'] else 'Unknown'
                    by_round.append({'round': round_name, 'count': r['count']})

                by_status = []
                completed_count = 0
                cancelled_count = 0
                rescheduled_count = 0
                for s in data["byStatus"]:
                    status_name = s['_id'] if s['_id'] else 'Unknown'
                    by_status.append({'status': status_name, 'count': s['count']})
                    if status_name == 'Completed':
                        completed_count = s['count']
                    elif status_name == 'Cancelled':
                        cancelled_count = s['count']
                    elif status_name == 'Rescheduled':
                        rescheduled_count = s['count']

                by_expert = []
                for e in data["byExpert"]:
                    expert_name = e['_id'] if e['_id'] else 'Unknown'
                    by_expert.append({'expert': expert_name, 'count': e['count']})

                interview_records = data["timeline"]
                completion_rate = round((completed_count / total) * 100, 1) if total > 0 else 0

                payload = {
                    "candidate_data": {
                        'name': candidate_name,
                        'total_interviews': total,
                        'completed_count': completed_count,
                        'cancelled_count': cancelled_count,
                        'rescheduled_count': rescheduled_count,
                        'completion_rate': completion_rate,
                        'by_round': by_round,
                        'by_status': by_status,
                        'by_expert': by_expert,
                        'unique_rounds': len(by_round),
                        'unique_experts': len(by_expert)
                    },
                    "interview_records": interview_records,
                }

            if cache:
                cache.set(cache_key, payload, timeout=300)
            cached = payload

        candidate_data = cached["candidate_data"]
        interview_records = cached["interview_records"]

    return render_template(
        'candidate_lookup.html',
        all_candidates=all_candidates,
        candidate_name=candidate_name,
        candidate_data=candidate_data,
        interview_records=interview_records
    )


def fetch_expert_activity_data(month_s, year_s, status_f, team_f=None, expert_f=None, include_all_candidates=False, exclude_rounds=None):
    """
    Shared function to fetch expert activity data.
    """
    cache = getattr(current_app, "cache", None)
    cache_key = expert_activity_cache_key(
        month_s,
        year_s,
        status_f,
        team_f,
        expert_f,
        include_all_candidates,
        exclude_rounds,
    )
    cached = cache.get(cache_key) if cache else None
    if cached is not None:
        displayed_experts = cached["displayed_experts"]
        po_counts_by_expert = build_po_counts_by_expert(month_s, year_s, displayed_experts)
        hydrated_team_data = []
        for team in cached["team_data"]:
            experts = []
            for expert in team["experts"]:
                expert_copy = dict(expert)
                expert_copy["po_count"] = po_counts_by_expert.get(
                    normalize_lookup_text(expert_copy["expert"]),
                    0,
                )
                experts.append(expert_copy)
            team_copy = dict(team)
            team_copy["experts"] = experts
            hydrated_team_data.append(team_copy)
        return hydrated_team_data, cached["summary"]

    db = get_db()
    management_directory = get_team_management_directory()
    reference = get_teams_reference()
    teams_map = reference["teams_map"]

    # Filter teams if team_f is provided
    if team_f and team_f in teams_map:
        teams_map_filtered = {team_f: teams_map[team_f]}
    else:
        teams_map_filtered = teams_map

    # Create expert to team mapping (using all teams to ensure correct mapping even if filtered)
    expert_to_team = {}
    for team_name, members in teams_map.items():
        for member in members:
            expert_to_team[normalize_lookup_text(member)] = team_name

    displayed_experts = []
    for members in teams_map_filtered.values():
        for expert in members:
            expert_clean = str(expert).strip()
            if expert_f and expert_clean != expert_f:
                continue
            if expert_clean:
                displayed_experts.append(expert_clean)

    # Pre-filter to the selected month using subject dates when available, while
    # keeping receivedDateTime as a safety net for malformed legacy subjects.
    subject_month_regex = build_selected_month_subject_regex(month_s, year_s)
    received_month_match = build_selected_month_received_date_match(month_s, year_s)
    match_query = {}
    month_filters = []
    if subject_month_regex:
        month_filters.append({"subject": {"$regex": subject_month_regex, "$options": "i"}})
    if received_month_match:
        month_filters.append({"receivedDateTime": received_month_match})
    if len(month_filters) == 1:
        match_query.update(month_filters[0])
    elif month_filters:
        match_query["$or"] = month_filters

    if status_f:
        match_query["status"] = status_f

    # Apply Exclude Rounds filter
    if exclude_rounds:
        if isinstance(exclude_rounds, str):
            rounds_to_exclude = [r.strip() for r in exclude_rounds.split(',') if r.strip()]
        else:
            rounds_to_exclude = exclude_rounds

        if rounds_to_exclude:
            # Use $nor to exclude documents where actualRound matches any of the patterns (case-insensitive)
            nor_conditions = [{"actualRound": {"$regex": r, "$options": "i"}} for r in rounds_to_exclude]
            if nor_conditions:
                 match_query["$nor"] = nor_conditions

    expert_patterns = [
        {"Expert": {"$regex": f"^{re.escape(str(expert).strip())}$", "$options": "i"}}
        for expert in displayed_experts
        if str(expert).strip()
    ]
    candidate_query = {
        "Candidate Name": {"$type": "string", "$ne": ""},
    }
    if expert_patterns:
        candidate_query["$or"] = expert_patterns
    else:
        candidate_query["Expert"] = "__no_match__"

    candidates = list(db.candidateDetails.find(
        candidate_query,
        {
            "Candidate Name": 1,
            "Expert": 1,
            "Recruiter": 1,
            "Branch": 1,
            "status": 1,
            "workflowStatus": 1,
            "_id": 0,
        }
    ))
    candidate_name_keys = {
        normalize_lookup_text(cand.get("Candidate Name"))
        for cand in candidates
        if normalize_lookup_text(cand.get("Candidate Name"))
    }

    # Query taskBody for interview details
    interview_pipeline = [
        {
            "$match": match_query
        },
        {
            "$project": {
                "CandidateKey": mongo_normalized_text("Candidate Name"),
                "Subject": "$subject",
                "ActualRound": "$actualRound",
                "Status": "$status",
                "Date": "$receivedDateTime",
            }
        },
        {
            "$match": {
                "CandidateKey": {"$in": list(candidate_name_keys) or ["__no_match__"]}
            }
        },
        {
            "$group": {
                "_id": "$CandidateKey",
                "InterviewCount": {"$sum": 1},
                "InterviewDetails": {"$push": {
                    "Subject": "$Subject",
                    "ActualRound": "$ActualRound",
                    "Status": "$Status",
                    "Date": "$Date"
                }}
            }
        }
    ]

    interview_results = list(db.taskBody.aggregate(interview_pipeline))
    selected_month_key = get_selected_po_month_key(month_s, year_s)
    candidate_interview_map = {}
    for result in interview_results:
        filtered_details = [
            detail
            for detail in result.get('InterviewDetails', [])
            if interview_matches_selected_month(detail, selected_month_key)
        ]
        candidate_interview_map[result['_id']] = {
            'count': len(filtered_details),
            'details': filtered_details,
        }

    # Build per_candidate and per_expert lists
    per_candidate = []
    expert_stats = defaultdict(lambda: {
        "ActiveCandidates": 0,
        "InactiveCandidates": 0,
        "TotalCandidates": 0,
        "TotalInterviews": 0,
        "StatusCounts": defaultdict(int)
    })
    active_candidates_by_expert = defaultdict(list)
    all_candidates_by_expert = defaultdict(list)

    def normalize_status(s):
        t = (s or "").strip().lower()
        if not t:
            return "(blank)"
        if t in {"active", "active candidate", "in process", "in-process", "inprogress", "completed", "needs_resume_understanding"}:
            return "Active"
        if t in {"backout", "back out", "backed out"}:
            return "Backout"
        if t in {"hold", "on hold"}:
            return "Hold"
        if t in {"low priority", "low-priority", "lowpriority", "low pri", "low"}:
            return "Low Priority"
        if t in {"placement offer", "placement offered", "offer", "offered", "placed"}:
            return "Placement Offer"
        return s.strip()

    for cand in candidates:
        cand_name = cand.get("Candidate Name", "")
        expert = cand.get("Expert", "")
        recruiter = cand.get("Recruiter", "")
        branch = cand.get("Branch", "")
        # Prioritize 'status' field as it contains the relevant values (Active, Backout, etc.)
        # Fallback to 'workflowStatus' if 'status' is missing
        raw_status = cand.get("status") or cand.get("workflowStatus")
        workflow_status = (raw_status or "").strip()
        expert_lower = normalize_lookup_text(expert)
        candidate_name_key = normalize_lookup_text(cand_name)

        interview_info = candidate_interview_map.get(candidate_name_key, {'count': 0, 'details': []})
        interview_count = interview_info['count']
        details = interview_info['details']

        # Sort details by date (newest first)
        details.sort(key=lambda x: x.get('Date', ''), reverse=True)

        # A candidate is active if they have interviews in the filtered period
        is_active = interview_count > 0

        per_candidate.append({
            "CandidateName": cand_name,
            "Expert": expert,
            "Recruiter": recruiter,
            "Branch": branch,
            "ExpertLower": expert_lower,
            "InterviewCount": interview_count,
            "InterviewDetails": details,
            "isActive": is_active
        })

        # Update expert stats
        expert_stats[expert_lower]["TotalCandidates"] += 1
        expert_stats[expert_lower]["TotalInterviews"] += interview_count
        status_key = normalize_status(workflow_status)
        expert_stats[expert_lower]["StatusCounts"][status_key] += 1

        cand_details = {
            "CandidateName": cand_name,
            "Recruiter": recruiter,
            "Branch": branch,
            "InterviewCount": interview_count,
            "InterviewDetails": details,
            "InterviewDate": details[0].get("Date", "") if details else "",
            "Subject": details[0].get("Subject", "") if details else "",
            "Status": details[0].get("Status", "") if details else "",
            "ActualRound": details[0].get("ActualRound", "") if details else "",
            "WorkflowStatus": workflow_status,
            "NormalizedStatus": status_key,
            "isActive": is_active
        }

        if is_active:
            expert_stats[expert_lower]["ActiveCandidates"] += 1
            active_candidates_by_expert[expert_lower].append(cand_details)
        else:
            expert_stats[expert_lower]["InactiveCandidates"] += 1

        if include_all_candidates:
            all_candidates_by_expert[expert_lower].append(cand_details)

    # Sort per_candidate
    per_candidate.sort(key=lambda x: (x["InterviewCount"], x["CandidateName"]), reverse=True)

    # Convert expert_stats to list format
    per_expert = [
        {"_id": expert, **stats}
        for expert, stats in expert_stats.items()
    ]
    per_expert.sort(key=lambda x: (x["ActiveCandidates"], x["TotalCandidates"]), reverse=True)

    # Map expert summaries (already created above)
    expert_summary = {normalize_lookup_text(e.get("_id")): e for e in per_expert}

    po_counts_by_expert = build_po_counts_by_expert(month_s, year_s, displayed_experts)

    # Build team_data using filtered teams map
    team_data = []
    for team_name, members in teams_map_filtered.items():
        team_active = team_inactive = team_total = 0
        expert_list = []

        for expert in members:
            # Clean expert name
            expert_clean = str(expert).strip()

            # Filter expert if expert_f is provided
            if expert_f and expert_clean != expert_f:
                continue

            key = normalize_lookup_text(expert_clean)
            e = expert_summary.get(key, {})
            active_cnt = e.get("ActiveCandidates", 0)
            inactive_cnt = e.get("InactiveCandidates", 0)
            total_cnt = e.get("TotalCandidates", 0)
            total_interviews = e.get("TotalInterviews", 0)

            team_active += active_cnt
            team_inactive += inactive_cnt
            team_total += total_cnt

            # Get active candidate list
            active_list = active_candidates_by_expert.get(key, [])
            active_list = sorted(
                active_list,
                key=lambda x: x.get("InterviewCount", 0),
                reverse=True
            )

            # Get all candidate list if requested
            all_list = []
            if include_all_candidates:
                all_list = all_candidates_by_expert.get(key, [])
                all_list = sorted(
                    all_list,
                    key=lambda x: (x.get("isActive", False), x.get("InterviewCount", 0)),
                    reverse=True
                )

            sc = e.get("StatusCounts", {})
            management = get_management_snapshot(expert_clean, directory=management_directory)
            expert_dict = {
                'expert': expert_clean,
                'expert_name': management['expert_name'] if management['expert_name'] != 'Unassigned' else (expert_clean.split('@')[0] if '@' in expert_clean else expert_clean),
                'total': total_cnt,
                'total_interviews': total_interviews,
                'active': active_cnt,
                'inactive': inactive_cnt,
                'team_name': team_name,
                'team_lead_name': management['team_lead_name'],
                'manager_name': management['manager_name'],
                'status_active': sc.get("Active", 0),
                'status_backout': sc.get("Backout", 0),
                'status_hold': sc.get("Hold", 0),
                'status_low_priority': sc.get("Low Priority", 0),
                'status_placement_offer': sc.get("Placement Offer", 0),
                'po_count': po_counts_by_expert.get(key, 0),
                'status_blank': sc.get("(blank)", 0),
                'grand_total': total_cnt,
                'active_candidates': active_list
            }

            if include_all_candidates:
                expert_dict['all_candidates'] = all_list

            expert_list.append(expert_dict)

        # Only add team if it has experts (after filtering)
        if expert_list:
            # Sort experts by active candidates
            expert_list.sort(key=lambda x: x['active'], reverse=True)

            team_data.append({
                'team': team_name,
                'total': team_total,
                'active': team_active,
                'inactive': team_inactive,
                'experts': expert_list
            })

    # Sort teams by active candidates
    team_data.sort(key=lambda x: x['active'], reverse=True)

    # Calculate overall stats
    overall_total = sum(t['total'] for t in team_data)
    overall_active = sum(t['active'] for t in team_data)
    overall_inactive = sum(t['inactive'] for t in team_data)
    overall_interviews = sum(e['total_interviews'] for t in team_data for e in t['experts'])

    summary = {
        'total_candidates': overall_total,
        'active_candidates': overall_active,
        'inactive_candidates': overall_inactive,
        'total_interviews': overall_interviews,
        'teams_count': len(team_data),
        'date_range': f"{month_s} {year_s}"
    }

    if cache:
        cache.set(
            cache_key,
            {
                "team_data": [
                    {
                        **team,
                        "experts": [{**expert, "po_count": 0} for expert in team["experts"]],
                    }
                    for team in team_data
                ],
                "summary": summary,
                "displayed_experts": displayed_experts,
            },
            timeout=300,
        )

    return team_data, summary


@candidates_bp.route('/expert-activity', methods=['GET'])
def expert_candidate_activity():
    """
    Expert Candidate Activity Dashboard
    Shows which experts have active candidates based on interview counts in a date range.

    Filters:
    - month: Month string (e.g., "JAN") for filtering by subject
    - year: Year string (e.g., "2024") for filtering by subject
    - status: Filter by interview status (optional)
    - team: Filter by team name (optional)
    - expert: Filter by expert email (optional)
    """
    # Date filters (Month/Year from Subject)
    current_date = datetime.utcnow()
    month = request.args.get('month', current_date.strftime('%b').upper())
    year = request.args.get('year', str(current_date.year))

    # Other filters
    status_filter = request.args.get('status', '')
    team_filter = request.args.get('team', '')
    expert_filter = request.args.get('expert', '')
    exclude_rounds = request.args.get('exclude_rounds', '')

    # Fetch data using shared function
    team_data, summary = fetch_expert_activity_data(month, year, status_filter, team_filter, expert_filter, exclude_rounds=exclude_rounds)
    po_access = get_current_po_access()
    po_security_enabled_for_counts = po_pin_security_enabled()
    po_counts_locked = po_security_enabled_for_counts and not po_access

    # Get all teams and experts for dropdowns (unfiltered)
    all_teams, all_experts, _ = get_team_options()

    # Generate list of years for filter (e.g., last 5 years)
    current_year = datetime.utcnow().year
    years = [str(y) for y in range(current_year, current_year - 5, -1)]
    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

    return render_template(
        'expert_activity.html',
        team_data=team_data,
        summary=summary,
        selected_month=month,
        selected_year=year,
        months=months,
        years=years,
        selected_status=status_filter,
        selected_team=team_filter,
        selected_expert=expert_filter,
        exclude_rounds=exclude_rounds,
        all_teams=all_teams,
        all_experts=all_experts,
        po_security_enabled=po_security_enabled_for_counts,
        po_access=po_access,
        po_counts_locked=po_counts_locked,
        po_unlock_url=url_for('po.po_access', next=current_request_next_url()) if po_counts_locked else ''
    )


@candidates_bp.route('/expert-activity/export', methods=['GET'])
def export_expert_activity():
    """Export expert candidate activity to Excel."""
    # Date filters (Month/Year from Subject)
    current_date = datetime.utcnow()
    month = request.args.get('month', current_date.strftime('%b').upper())
    year = request.args.get('year', str(current_date.year))

    # Other filters
    status_filter = request.args.get('status', '')
    team_filter = request.args.get('team', '')
    expert_filter = request.args.get('expert', '')
    exclude_rounds = request.args.get('exclude_rounds', '')
    po_counts_locked = po_pin_security_enabled() and not get_current_po_access()

    # Fetch data using shared function
    team_data, _ = fetch_expert_activity_data(
        month, year, status_filter, team_filter, expert_filter, include_all_candidates=True, exclude_rounds=exclude_rounds
    )

    # Build Excel data
    summary_rows = []
    detail_rows = []

    for team in team_data:
        team_name = team['team']
        for expert_data in team['experts']:
            expert_name = expert_data['expert']

            # Summary Row
            summary_rows.append({
                "Team": team_name,
                "Expert": expert_name,
                "TotalCandidates": expert_data['total'],
                "ActiveCandidates": expert_data['active'],
                "InactiveCandidates": expert_data['inactive'],
                "TotalInterviews": expert_data['total_interviews'],
                "Active": expert_data['status_active'],
                "Backout": expert_data['status_backout'],
                "Hold": expert_data['status_hold'],
                "Low Priority": expert_data['status_low_priority'],
                "Placement Offer": "Locked" if po_counts_locked else expert_data.get('po_count', 0),
                "(blank)": expert_data['status_blank'],
                "Grand Total": expert_data['grand_total']
            })

            # Detail Rows (using all_candidates which is populated because include_all_candidates=True)
            candidates_list = expert_data.get('all_candidates', [])

            for c in candidates_list:
                interview_details = c.get("InterviewDetails", [])
                if interview_details:
                    for d in interview_details:
                        detail_rows.append({
                            "Team": team_name,
                            "Expert": expert_name,
                            "CandidateName": c.get("CandidateName", ""),
                            "Recruiter": c.get("Recruiter", ""),
                            "Branch": c.get("Branch", ""),
                            "InterviewDate": d.get("Date", ""),
                            "Subject": d.get("Subject", ""),
                            "Status": d.get("Status", ""),
                            "ActualRound": d.get("ActualRound", "")
                        })
                else:
                    # Candidate with no interviews in period
                    detail_rows.append({
                        "Team": team_name,
                        "Expert": expert_name,
                        "CandidateName": c.get("CandidateName", ""),
                        "Recruiter": c.get("Recruiter", ""),
                        "Branch": c.get("Branch", ""),
                        "InterviewDate": "",
                        "Subject": "",
                        "Status": "",
                        "ActualRound": ""
                    })

    # Create DataFrames
    summary_df = pd.DataFrame(summary_rows)
    detail_df = pd.DataFrame(detail_rows)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        else:
             pd.DataFrame(columns=["Team", "Expert", "TotalCandidates"]).to_excel(writer, sheet_name='Summary', index=False)

        if not detail_df.empty:
            # Sort for better grouping in sheet
            sort_cols = ["Team", "Expert", "CandidateName", "Recruiter", "Branch", "InterviewDate"]
            for col in sort_cols:
                if col not in detail_df.columns:
                    detail_df[col] = ""
            detail_df_sorted = detail_df.sort_values(by=sort_cols, na_position='last')
            detail_df_sorted.to_excel(writer, sheet_name='CandidateDetails', index=False)

            # Merge repeated cells hierarchically
            ws = writer.sheets['CandidateDetails']
            # Map column names to 1-based indices
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_index_map = {name: idx + 1 for idx, name in enumerate(header)}

            # Define merge hierarchy: Each group includes parent columns to ensure we don't merge across boundaries
            merge_groups = [
                ["Team"],
                ["Team", "Expert"],
                ["Team", "Expert", "CandidateName"],
                ["Team", "Expert", "CandidateName", "Recruiter"],
                ["Team", "Expert", "CandidateName", "Branch"]
            ]

            for group_cols in merge_groups:
                target_col_name = group_cols[-1]
                if target_col_name not in col_index_map:
                    continue

                col_idx = col_index_map[target_col_name]

                # Group by hierarchy and iterate
                # sort=False preserves the original sorted order of rows
                grouped = detail_df_sorted.groupby(group_cols, dropna=False, sort=False)

                start_row = 2  # Data starts at row 2 (1-based index)
                for _, df_group in grouped:
                    length = len(df_group)
                    end_row = start_row + length - 1

                    if length > 1:
                        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

                        # Center align merged cells vertically
                        cell = ws.cell(row=start_row, column=col_idx)
                        cell.alignment = Alignment(vertical='center', horizontal='left')

                    start_row = end_row + 1
        else:
             pd.DataFrame(columns=["Team", "Expert", "CandidateName"]).to_excel(writer, sheet_name='CandidateDetails', index=False)

    output.seek(0)
    filename = f"expert_candidate_activity_{month}_{year}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@candidates_bp.route('/active', methods=['GET'])
def active_candidates():
    return redirect(url_for('dashboard.index'))
